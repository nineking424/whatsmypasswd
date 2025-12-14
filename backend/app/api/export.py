import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import Workbook, load_workbook

from app.db.database import get_db
from app.api.auth import verify_token
from app.models import Credential, Category, AuditLog, AuditAction, CredentialType
from app.services.crypto import get_crypto_service

router = APIRouter(prefix="/export", tags=["export"])
crypto = get_crypto_service()


@router.get("/excel")
async def export_to_excel(
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(verify_token),
):
    """Export all credentials to Excel file."""
    query = (
        select(Credential)
        .options(selectinload(Credential.category))
        .order_by(Credential.type, Credential.name)
    )
    result = await db.execute(query)
    credentials = result.scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    # Header
    headers = [
        "Type", "Name", "Host", "Port", "Username", "Password",
        "Category", "Tags", "Description", "Extra Data"
    ]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for cred in credentials:
        host = crypto.decrypt(cred.host) if cred.host else ""
        username = crypto.decrypt(cred.username) if cred.username else ""
        password = crypto.decrypt(cred.password) if cred.password else ""
        extra_data = crypto.decrypt_dict(cred.extra_data) if cred.extra_data else {}

        ws.append([
            cred.type.value,
            cred.name,
            host,
            cred.port or "",
            username,
            password,
            cred.category.name if cred.category else "",
            ",".join(cred.tags or []),
            cred.description or "",
            str(extra_data) if extra_data else "",
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=credentials.xlsx"},
    )


@router.post("/excel")
async def import_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(verify_token),
):
    """Import credentials from Excel file."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload an Excel file (.xlsx or .xls)",
        )

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active

        # Get existing categories
        category_query = select(Category)
        category_result = await db.execute(category_query)
        categories = {c.name.lower(): c for c in category_result.scalars().all()}

        imported = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Skip empty rows
                continue

            try:
                type_str, name, host, port, username, password, category_name, tags_str, description, extra_data_str = (
                    row + (None,) * (10 - len(row))
                )[:10]

                if not type_str or not name:
                    errors.append(f"Row {row_num}: Type and Name are required")
                    continue

                # Validate type
                try:
                    cred_type = CredentialType(type_str.lower())
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid type '{type_str}'")
                    continue

                # Get or create category
                category_id = None
                if category_name:
                    category_key = category_name.lower()
                    if category_key in categories:
                        category_id = categories[category_key].id
                    else:
                        new_category = Category(name=category_name)
                        db.add(new_category)
                        await db.flush()
                        categories[category_key] = new_category
                        category_id = new_category.id

                # Parse tags
                tags = [t.strip() for t in tags_str.split(",")] if tags_str else []

                # Parse extra_data
                extra_data = None
                if extra_data_str:
                    try:
                        import ast
                        extra_data = ast.literal_eval(extra_data_str)
                    except:
                        pass

                # Create credential
                credential = Credential(
                    type=cred_type,
                    name=name,
                    host=crypto.encrypt(host) if host else None,
                    port=int(port) if port else None,
                    username=crypto.encrypt(username) if username else None,
                    password=crypto.encrypt(password) if password else None,
                    category_id=category_id,
                    tags=tags,
                    description=description,
                    extra_data=crypto.encrypt_dict(extra_data) if extra_data else None,
                )
                db.add(credential)
                imported += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        await db.flush()

        # Log import action
        log = AuditLog(
            action=AuditAction.CREATE,
            credential_name=f"Excel import: {imported} items",
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent", "")[:255],
        )
        db.add(log)

        return {
            "imported": imported,
            "errors": errors[:10],  # Return first 10 errors
            "total_errors": len(errors),
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}",
        )
